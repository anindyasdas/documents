"""
/*-------------------------------------------------
 * Copyright(c) 2020 by LG Electronics.
 * Confidential and Proprietary All Rights Reserved.
 *-------------------------------------------------*/
@author: vishwaas@lge.com
"""
import collections
import numpy as np
from . import constants
import pandas as pd
import shutil
import scipy
from sklearn.model_selection import StratifiedShuffleSplit

from sklearn import metrics


def append_df_to_excel(filename, df, sheet_name,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass
    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=0, **to_excel_kwargs)

    # save the workbook
    writer.save()


# Number of vertices

def __dfs(mat, marked, vert, start, paths, path, vals):
    """
    Does DFS on the matrix to find all cycles of k nodes
    :param mat: cosine matrix
    :param marked: boolean flag to check if the node is visited
    :param vert: node
    :param start: start node of the cycle
    :param paths: best path
    :param path: current path
    :param: vals: list of [n: counter, v: number of vertices, curr_score: current score]
    """
    marked[vert] = True

    if vals[0] == 0:

        marked[vert] = False

        if len(paths) == 0:
            paths.append((path, vals[2]))
        else:
            prev_path, prev_score = paths[0]
            if vals[2] > prev_score:
                paths[0] = (path, vals[2])
        return
    for i in range(vals[1]):
        if not marked[i]:
            next_path = path[:]
            next_path.append(i)
            next_score = __cum_score(mat, next_path, vals[2], i)
            __dfs(mat, marked, i, start, paths, next_path, [vals[0] - 1, vals[1], next_score])

    marked[vert] = False
    return


def __cum_score(mat, paths, curr_score, ind):
    """
    Gets the updated score after adding a new question
    :param mat: cosine matrix
    :param paths: indices of questions in the current set
    :param curr_score: current score
    :param ind: new index
    :return: new score
    """
    next_score = curr_score
    for i in paths:
        next_score += mat[ind, i]
    return next_score


def get_top_questions(questions, k, siam_model):
    """
    Gets the top k canonical questions by choosing top k questions that have the maximum cosine distance between each other
    :param questions: list of str
    :param k: int
    :param siam_model: siamese bert model
    :return: top k questions
    """
    questions = sorted(questions, key=len)
    if k == 1:
        ind = np.argmin([len(q) for q in questions])
        return [questions[ind]]

    if k == 2:
        max_ind = (0, 1)
        max_val = -1
        mat = __get_sim_mat(questions, siam_model)
        for i in range(mat.shape[0]):
            for j in range(i + 1, mat.shape[1]):
                if mat[i, j] > max_val:
                    max_val = mat[i, j]
                    max_ind = (i, j)
        return [questions[max_ind[0]], questions[max_ind[1]]]
    else:
        if len(questions) >= constants.CANONICAL_QUESTION_PRUN:
            out = __get_break_sim(questions, siam_model, k)
            return [questions[i] for i in out]
        else:
            mat = __get_sim_mat(questions, siam_model)
            inds = __get_top_questions_k3(mat, k)[0]
            return [questions[i] for i in inds]


def __get_break_sim(questions, siam_model, k):
    """
    Since the complexity is high for a high len(questions), we break it into smaller chunks, although this might reduce the accuracy
    :param questions: list of str
    :param k: int
    :param siam_model: siamese bert model
    :return: top k questions
    """
    div = int(np.ceil(len(questions) / constants.CANONICAL_QUESTION_PRUN))
    out = []
    m = int(np.ceil(k / div))
    for i in range(div):
        ind_div = np.arange(i, len(questions), div)
        ques_div = [questions[j] for j in ind_div]
        mat_div = __get_sim_mat(ques_div, siam_model)
        out_div = __get_top_questions_k3(mat_div, m)[0]
        out.extend(out_div)

    return out


def __get_sim_mat(questions, siam_model):
    """
    Gets 2d matrix of cosine similarity between all questions will all other questions
    :param questions: list of str
    :param siam_model: siamese bert model
    :return: 2d matrix
    """
    embeddings = siam_model.compute_embeddings_single(questions)
    mat = np.zeros([len(questions), len(questions)])
    for i, s1 in enumerate(embeddings):
        for j, s2 in enumerate(embeddings):
            mat[i, j] = scipy.spatial.distance.cosine(s1, s2)

    return mat


def __get_top_questions_k3(mat, k):
    """
    Does DFS on the matrix to find all cycles of k nodes
    :param mat: 2d matrix
    :param k: int
    :return: top indices
    """
    V = mat.shape[0]
    marked = [False] * V

    paths = []

    for i in range(V - (k - 1)):
        __dfs(mat, marked, i, i, paths, [i], [k - 1, V, mat[i][i]])

        marked[i] = True

    return paths[0]


def split_train_test():
    """
    Splits the data into train and test files in a stratified manner
    :return: None
    """
    full_data = []
    for prod in [constants.WASHING_MACHINE, constants.VACUUM_CLEANER, constants.MICROWAVE_OVEN]:
        data = pd.read_excel(constants.INPUT_FILES[prod][constants.SPEC], sheet_name=constants.L1_L2_L3)
        data['product'] = [prod] * data.shape[0]
        full_data.append(data[['Questions', 'Key', 'product']])
    full_data = pd.concat(full_data, axis=0)

    full_data['Reason/Solution'] = full_data['Key']
    full_data['User Question'] = full_data['Questions']
    full_data['prod_key'] = full_data[['Key', 'product']].apply(lambda x: '|'.join(x), axis=1)
    sss = StratifiedShuffleSplit(n_splits=1, test_size=0.4, random_state=0)
    sss_split = sss.split(full_data['User Question'].values, full_data['prod_key'].values)
    train_data, test_data = next(sss_split)
    full_data = full_data.drop(['Key', 'Questions', 'prod_key'], axis=1)

    train_data = full_data.iloc[train_data]
    test_data = full_data.iloc[test_data]
    train_data.to_excel(constants.CUSTOM_SPEC_TRAIN_DATA, index=False, sheet_name='sheet1')
    test_data.to_excel(constants.CUSTOM_SPEC_TEST_DATA, index=False, sheet_name='sheet1')


if __name__ == '__main__':
    questions = ['Electron is a particle', 'Proton is a particle',
                 'An Electron is extremely small', 'One of the particles is an electron', 'Football is a sport',
                 'Barack Obama was a president']
    k = 3
    from components.text_similarity.siamese_bert import SiameseBERT
    import scipy

    siam = SiameseBERT(None)
    print(get_top_questions(questions, k, siam))