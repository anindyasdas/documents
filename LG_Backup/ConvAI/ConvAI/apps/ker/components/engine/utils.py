"""
/*-------------------------------------------------
* Copyright(c) 2020 by LG Electronics.
* Confidential and Proprietary All Rights Reserved.
*-------------------------------------------------*/
@author: vishwaas.n@lge.com
"""
import pandas as pd
from collections import defaultdict
import re
from sklearn.model_selection import StratifiedShuffleSplit


class Utils:

    @staticmethod
    def append_df_to_excel(filename, df, sheet_name,
                           truncate_sheet=False,
                           **to_excel_kwargs):
        """
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.

        Parameters:
          filename : File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
          df : dataframe to save to workbook
          sheet_name : Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
          startrow : upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
          truncate_sheet : truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
          to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                            [can be dictionary]

        Returns: None
        """
        from openpyxl import load_workbook

        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(filename, engine='openpyxl')

        try:
            # try to open an existing workbook
            writer.book = load_workbook(filename)

            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)

            # copy existing sheets
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            # file does not exist yet, we will create it
            pass
        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=0, **to_excel_kwargs)

        # save the workbook
        writer.save()

    @staticmethod
    def split_train_test(input_file, stratify_on: list, sheet_name='L1+L2+L3'):
        """
        Splits the data into train and test files in a stratified manner
        :param input_file: input file path
        :param stratify_on: list of column names to stratify on
        :param sheet_name: sheet name of the data
        :return: train and test data
        """
        strat_key = 'strat_key'
        cnt_key = 'cnt'
        full_data = pd.read_excel(input_file, sheet_name=sheet_name, engine='openpyxl')
        for col in full_data.columns:
            full_data[col] = full_data[col].apply(Utils.pre_process_key)
        full_data[strat_key] = full_data[stratify_on].apply(lambda x: '|'.join(x), axis=1)
        cnt = defaultdict(int)
        for k in full_data[strat_key].values:
            cnt[k] += 1
        full_data[cnt_key] = full_data[strat_key].apply(lambda x: cnt[x])
        extra_data = full_data[full_data[cnt_key] == 1]
        full_data = full_data[full_data[cnt_key] != 1]
        sss = StratifiedShuffleSplit(n_splits=1, test_size=0.3, random_state=0)
        sss_split = sss.split(full_data, full_data[strat_key].values)
        train_data, test_data = next(sss_split)

        train_data = full_data.iloc[train_data]
        test_data = full_data.iloc[test_data]
        train_data = pd.concat([train_data, extra_data], axis=0)
        train_data = train_data.drop([cnt_key, strat_key], axis=1)
        test_data = test_data.drop([cnt_key, strat_key], axis=1)

        return train_data, test_data

    @staticmethod
    def pre_process_key(key):
        """
        preprocesses the data by applying lower and removing trailing fullstops and spaces
        :param key: str - key
        :return: pre processed key
        """
        new_key = str(key)
        new_key = new_key.lower()
        new_key = new_key.strip()
        new_key = re.sub('\s+', ' ', new_key)
        new_key = re.sub('\.$', '', new_key)
        return new_key


if '__main__' == __name__:
    Utils.split_train_test('WM+REF+DISHWASHER_L1+L2+L3_SPEC.xlsx', ['Key', 'Sub key'])